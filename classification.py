import os
import cv2
import face_recognition
import numpy as np
import threading
import time
import csv
import openpyxl
from datetime import datetime
from tkinter import Tk, messagebox

# Direktori penyimpanan gambar wajah
faces_folder = "faces/"
csv_file = "faces_list.csv"
excel_file = "absensi.xlsx"

# List untuk menyimpan wajah dan nama yang dikenali
known_faces = []
known_names = []

# Baca data dari CSV jika tersedia
if os.path.exists(csv_file):
    with open(csv_file, mode="r", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        for row in reader:
            img_path = os.path.join(faces_folder, row["gambar"])
            if os.path.exists(img_path):
                img = face_recognition.load_image_file(img_path)
                encoding = face_recognition.face_encodings(img)
                if encoding:
                    known_faces.append(encoding[0])
                    known_names.append(row["nama"])  # Simpan nama sesuai indeks

print(f"\nTotal wajah dikenali: {len(known_faces)}")

# Inisialisasi kamera
camera = cv2.VideoCapture(0, cv2.CAP_DSHOW)

# Variabel global
progress = 0
progress_lock = threading.Lock()
current_name = "Unknown"

# Fungsi untuk menyimpan data ke Excel
def save_to_excel(name):
    if not os.path.exists(excel_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Nama", "Tanggal", "Waktu"])  # Tambahkan header
        workbook.save(excel_file)

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # Cek apakah pengguna sudah absen hari ini
    for row in sheet.iter_rows(values_only=True):
        if row[0] == name and row[1] == date_str:
            messagebox.showwarning("Peringatan", f"{name} sudah absen hari ini!")
            return

    # Tambahkan data baru
    sheet.append([name, date_str, time_str])
    workbook.save(excel_file)

# Fungsi untuk mengenali wajah
def recognize_faces(frame):
    global current_name

    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

    recognized = False
    detected_name = "Unknown"

    for encoding in face_encodings:
        matches = face_recognition.compare_faces(known_faces, encoding, 0.6)
        if True in matches:
            face_distances = face_recognition.face_distance(known_faces, encoding)
            best_match_index = np.argmin(face_distances)

            if matches[best_match_index]:
                detected_name = known_names[best_match_index]
                recognized = True

    if recognized:
        current_name = detected_name

    return face_locations, detected_name, recognized

# Fungsi untuk menggambar kotak wajah dan progress bar
def draw_boxes(frame):
    global progress

    face_locations, name, recognized = recognize_faces(frame)

    for (top, right, bottom, left) in face_locations:
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 3)
        cv2.putText(frame, name, (left, bottom + 30), cv2.FONT_HERSHEY_SIMPLEX, 
                    0.8, (255, 255, 255), 2, cv2.LINE_AA)

    with progress_lock:
        if recognized:
            progress = min(progress + 10, 100)
        else:
            progress = max(progress - 2, 0)

    cv2.rectangle(frame, (50, 400), (550, 430), (255, 255, 255), 2)
    cv2.rectangle(frame, (50, 400), (50 + int(progress * 5), 430), (0, 255, 0), -1)
    cv2.putText(frame, f"Progress: {progress}%", (250, 395), cv2.FONT_HERSHEY_SIMPLEX, 
                0.7, (255, 255, 255), 2)

# Menampilkan pop-up dan menyimpan data absensi
def show_popup():
    global current_name
    root = Tk()
    root.withdraw()
    messagebox.showinfo("Absensi", f"Berhasil Absen!\nNama: {current_name}")
    root.destroy()
    save_to_excel(current_name)
    close_window()

# Cek progres absen
def check_progress():
    global progress
    while True:
        if progress == 100:
            show_popup()
            time.sleep(2)
            with progress_lock:
                progress = 0
        time.sleep(0.5)

# Menutup jendela
def close_window():
    camera.release()
    cv2.destroyAllWindows()
    exit()

# Main loop
def main():
    threading.Thread(target=check_progress, daemon=True).start()

    while True:
        ret, frame = camera.read()
        draw_boxes(frame)
        cv2.imshow('Face Recognition', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            close_window()

if __name__ == '__main__':
    main()
